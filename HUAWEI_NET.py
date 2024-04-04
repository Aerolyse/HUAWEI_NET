import urllib3
import json
import re
import sys
import argparse
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill ,Font
from datetime import datetime


REQUEST_HTTPS= "https://weu.naas.huawei.com:18002/controller"

def get_token():
    https = urllib3.PoolManager()
    headers = {
        'Content-Type': 'application/json'
        # Add other headers here
    }
    data = {
        "userName": "user",
        "password": "password"
    }
    encoded_data = json.dumps(data).encode('utf-8')
    response = https.request(
        'POST',
        'https://weu.naas.huawei.com:18002/controller/v2/tokens',
        headers=headers,
        body=encoded_data
    )
    response = json.loads(response.data.decode('utf-8'))
    
    if response["errcode"] == "0":
        token_id_value =  response['data']["token_id"]
        return token_id_value
    else :
        print(token_id_value["errmsg"])

def get_device(token):
        
        https = urllib3.PoolManager()
        headers = {
            'Content-Type': 'application/json',
            'X-ACCESS-TOKEN': token
        } 
        response = https.request(
            'GET',
            'https://weu.naas.huawei.com:18002/controller/campus/v3/devices',
            headers=headers,
        )
        device_dict = json.loads(response.data.decode('utf-8'))
        if device_dict['errcode'] == "0":
            return device_dict
        else:
            print("ERROR ON THE get_device() function")
            sys.exit(1) 
        
def Device_Sorter(site_Name_args,device_dict,device_Type_args="LSW"):
    devices_List = []
    for i in device_dict["data"]:    
        device_type = i["deviceType"]
        site_Name=i["siteName"]
        device_Name = i["name"]
        esn=i["esn"]
        device_Id=i["id"]

        if device_type == device_Type_args and site_Name == site_Name_args:
            devices_List.append([device_Name,device_Id,esn,site_Name])
    return devices_List

def export_xlsx(token,site_devices_List,custom_Name_Path):
    date_time = str(datetime.now()).replace("-", "_").replace(" ", "_").replace(":", "_").replace(".", "_")
    site_Name = site_devices_List[0][3]
    wb = Workbook()
    sheet = wb.active
    sheet.title = site_Name
    sheet.freeze_panes = "A2"
    actual_row = 2
    already_seen_devices = []
    ethernet_first_row = False
    ethTrunk_first_row = False
    grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    for i in site_devices_List:
        switch_number_increment_list = []
        switch_number = 0
        device_Id = i[1]
        device_Name = i[0]
        https = urllib3.PoolManager()
        headers = {
            'Content-Type': 'application/json',
            'X-ACCESS-TOKEN': token
        }
        
        data = {
                "deviceIdList":[device_Id]   
        }
        encoded_data = json.dumps(data).encode('utf-8')
        response = https.request(
            'POST',
            'https://weu.naas.huawei.com:18002/controller/campus/v3/stack/ids/list',
            headers=headers,
            body=encoded_data
        )
        device_stack = json.loads(response.data.decode('utf-8'))
        stackId = device_stack['data'][0]["stackId"]
        

        if len(stackId) == 0:
            request_Id = device_Id
            stackId == "no stack"
            print("standalone:",request_Id)
        else :
            request_Id = stackId
            print("stackid:",request_Id)        
        
        if request_Id in already_seen_devices:
                print("continue")
                continue
        elif request_Id not in already_seen_devices:
                print("appended")
                already_seen_devices.append(request_Id)
        print(already_seen_devices)
        https = urllib3.PoolManager()
        headers = {
            'Content-Type': 'application/json',
            'X-ACCESS-TOKEN': token
        } 
        response = https.request(
            'GET',
            f"https://weu.naas.huawei.com:18002/controller/campus/v1/networkservice/networkconfig/net/lswport/devices/{request_Id}/ports",
            headers=headers,
        )
        port_dict = json.loads(response.data.decode('utf-8'))
        ethernet_List = port_dict['data'][0]["ethernetList"]
        ethTrunk_List = port_dict['data'][0]["ethTrunkList"]
        if ethernet_first_row and ethTrunk_first_row:
            pass
        else :
            if len(ethernet_List) > 0:
                ethernet_first_row = True
                ethernet_keys = [key for key,value in ethernet_List[0].items()]
                First_row_keys = ethernet_keys
                First_row_keys.insert(0,"device_Name")
                First_row_keys.insert(1,"interface_Type")
                First_row_keys.extend(["device_Id","stack_Id"])
                First_row_keys.remove("name")
                First_row_keys.insert(3,"name")
            #if len(ethernet_List) == 0: stop script no interface detected 
            if not ethTrunk_first_row and ethernet_first_row and len(ethTrunk_List) > 0:
                ethTrunk_first_row = True
                ethTrunk_keys = [key for key, value in ethTrunk_List[0].items()] + [dict_key for key, value in ethTrunk_List[0].items() if isinstance(value, dict) for dict_key, dict_value in value.items()] 
                First_row_keys = ethernet_keys.copy() + [key for key in ethTrunk_keys if key not in ethernet_keys]
                First_row_keys.remove("portMemberList")
                First_row_keys.insert(90,"portMemberList")
                First_row_keys.remove("device_Id")
                First_row_keys.remove("stack_Id")
                First_row_keys.extend(["device_Id","stack_Id"])

        # Write First_row_keys to the first row
        for idx, key in enumerate(First_row_keys, start=1):
            sheet.cell(row=1, column=idx, value=key)

        # Write the values to the corresponding columns
        for row_index, interface_data_dict in enumerate(ethernet_List, start=actual_row):
            # Iterate over each key in First_row_keys
            interface_Name = interface_data_dict["name"]
            pattern = r'\d+/\d+/\d+'
            match = re.search(pattern,interface_Name)
            switch_number_increment = int(match.group()[0])
            if switch_number_increment in switch_number_increment_list:
                pass
            if switch_number_increment not in switch_number_increment_list:
                switch_number_increment_list.append(switch_number_increment)
                switch_number+=1
                
            print(switch_number_increment_list)
            for col_index, key in enumerate(First_row_keys, start=1):
                # Access the cell in the Excel sheet
                cell = sheet.cell(row=row_index, column=col_index)
                
                # Retrieve the value corresponding to the key
                value = interface_data_dict.get(key)
                
                
                if key == "device_Name":
                    cell.value = device_Name + f"_SW{switch_number}"
                elif key == "interface_Type":
                        cell.value = "Ethernet"
                elif key == "device_Id":
                    cell.value = device_Id
                elif key == "stack_Id":
                    cell.value = stackId
                elif value is None and key in interface_data_dict:
                    cell.value = "None"
                else:
                    cell.value = value if value is not None else " "
        actual_row = actual_row + len(ethernet_List)

        if len(ethTrunk_List) > 0:
            for row_index, interface_data_dict in enumerate(ethTrunk_List, start=actual_row):
                # Iterate over each key in First_row_keys
                for col_index, key in enumerate(First_row_keys, start=1):
                    # Access the cell in the Excel sheet
                    cell = sheet.cell(row=row_index, column=col_index)
                    # Retrieve the value corresponding to the key
                    value = interface_data_dict.get(key)

                    if key == "device_Name":
                        cell.value = device_Name + f"_SW{switch_number}"
                    elif key == "interface_Type":
                        cell.value = "Eth-Trunk"
                    elif key == "device_Id":
                        cell.value = device_Id
                    elif key == "stack_Id":
                        cell.value = stackId
                    elif value is None and key in interface_data_dict:
                            cell.value = "None"
                    
                    elif isinstance(value, list):
                        # Join the list items with a comma and set the cell value
                        cell.value = ', '.join(value)
                    # Check if the value is a dictionary
                    elif isinstance(value, dict):
                        # Iterate over the sub-keys and values in the nested dictionary
                        for sub_key, sub_value in value.items():
                            # Check if the sub-key is in First_row_keys
                            if sub_key in First_row_keys:
                                # Find the column index for the sub-key
                                sub_col_index = First_row_keys.index(sub_key) + 1
                                # Access the cell in the Excel sheet for the sub-key
                                sub_cell = sheet.cell(row=row_index, column=sub_col_index)
                                # Set the cell value for the sub-key
                                if sub_value is None:
                                    sub_cell.value = 'None'
                                else:
                                    sub_cell.value = sub_value
                                    
                    else:
                # Set the cell value for non-dictionary values if the cell is empty or the value is not None
                        if value is not None or len(str(cell.value)) == 0:
                            if value is not None:
                                cell.value = value
                        if key not in interface_data_dict and key not in interface_data_dict["LswEthTrunkLacpConfigDto"] :
                            cell.value = " "
                        
            actual_row = actual_row + len(ethTrunk_List) 
        

    for col in sheet.iter_cols():
        if col[0].value != 'description':
            for cell in col:
                if cell.value == " ":
                    cell.fill = grey_fill  

    for cell_bold_first_row in sheet["1"]:
        cell_bold_first_row.font = Font(bold=True)
    for cell_bold_first_collumn in sheet["A"]:
        cell_bold_first_collumn.font = Font(bold=True)

# Freeze the first row
    wb.save(f"{custom_Name_Path}")
            

def update_ethernet_request(token,body_dict,Id,Device_name,ethernet_interface_Name):
    # Remove extra double quotes from description value
    https = urllib3.PoolManager()
    request_headers = {
        'Content-Type': 'application/json',
        'X-AUTH-TOKEN': token
        # Add other headers here
    }
    #print("BODY ------->",body_dict)
    
    # Encapsulate the body dictionary within a list
    json_data = json.dumps([body_dict]).encode('utf-8')
    response = https.request(
        "PUT",
        f"{REQUEST_HTTPS}/campus/v1/networkservice/networkconfig/net/lswport/devices/{Id}/ethernet-ports",
        headers=request_headers,
        body=json_data
    )
    
    response = json.loads(response.data.decode('utf-8'))
    if response["errcode"] == "0":
        print(f"✅ The {ethernet_interface_Name} modification as been succesfully appliedon the switch:{Device_name}")
    else:
        print(f"❌ An error occured while applying the modification \nInterface: {ethernet_interface_Name}\nSwitch:{Device_name}\n{response}")

def update_ethtrunk_request(token,body,Id,Device_name,ethtrunk_name):
    # Remove extra double quotes from description value
    https = urllib3.PoolManager()
    request_headers = {
        'Content-Type': 'application/json',
        'X-AUTH-TOKEN': token
        # Add other headers here
    }
    
    # Encapsulate the body dictionary within a list
    json_data = json.dumps(body).encode('utf-8')
    response = https.request(
        "PUT",
        f"{REQUEST_HTTPS}/campus/v1/networkservice/networkconfig/net/lswport/devices/{Id}/ethtrunk-ports/{ethtrunk_name}",
        headers=request_headers,
        body=json_data
    )
    
    response = json.loads(response.data.decode('utf-8'))
    if response["errcode"] == "0":
        print(f"✅ The {ethtrunk_name} modification as been succesfully appliedon the switch:{Device_name}\n {response}")
    else:
        print(f"❌ An error occured while applying the modification \nInterface: {ethtrunk_name}\nSwitch:{Device_name}")
     
def get_interfaces(token,Id,interface_Type):
    # Remove extra double quotes from description value
    https = urllib3.PoolManager()
    request_headers = {
        'Content-Type': 'application/json',
        'X-AUTH-TOKEN': token
        # Add other headers here
    }
    response = https.request(
        "GET",
        f"{REQUEST_HTTPS}/campus/v1/networkservice/networkconfig/net/lswport/devices/{Id}/ports",
        headers=request_headers,
    )
    
    if interface_Type == "Ethernet":
        ethernet_interfaces = json.loads(response.data.decode('utf-8'))
        ethernet_interfaces = ethernet_interfaces["data"][0]["ethernetList"]
        return ethernet_interfaces
    
    elif interface_Type == "Eth-Trunk":
        ethtrunk_interfaces = json.loads(response.data.decode('utf-8'))
        ethtrunk_interfaces = ethtrunk_interfaces["data"][0]["ethTrunkList"]
        return ethtrunk_interfaces
    
def dict_comparator(compared_dict,comparator_dict,interface_Type):
    ######################################################################
    # THIS IS A SECURITY IN CASE OF DURING THE CONVERSION TO DICT        #  
    # A NON DESIRED KEY SLIPPED THROUGH THE DICT THAT WILL BE CONVERTED  #
    # TO JSON THEN SENT TO THE API                                       #
    ######################################################################
    ethtrunk_keys_to_delete = []
    if interface_Type == "Ethernet":
        try:
            keys_to_delete = set(compared_dict.keys()) - set(comparator_dict[0].keys())
            # Delete the keys from ompared_dict
            for key in keys_to_delete:
                del compared_dict[key]
            return(compared_dict)
        except Exception as e:
            print(f"ERROR WITH THE FUNCTION DICT_COMPARATOR:\n {e}")
            sys.exit(1)
    elif interface_Type == "Eth-Trunk":
        try:
            keys_to_delete = set(compared_dict.keys()) - set(comparator_dict[0].keys())
            for key in keys_to_delete:
                del compared_dict[key]
            
            # Iterate through the keys and values of the compared_dict
            for key, value in compared_dict.items():
    # If the value is a dictionary and the key exists in comparator_dict[0]
                    # If the value is a dictionary and the key exists in comparator_dict[0]
                if isinstance(value, dict) and key in comparator_dict[0]:
                    # Recursively compare the nested dictionaries
                    nested_comparator = comparator_dict[0][key]
                    dict_comparator(value, [nested_comparator], interface_Type)
                else:
                    # Print the key before deleting it
                    #print(f"Deleting key: {key}")
                    # Add the key to the list of keys to delete
                    ethtrunk_keys_to_delete.append(key)
                                
            return compared_dict
        except Exception as e:
            print(f"ERROR WITH THE FUNCTION DICT_COMPARATOR :\n {e}")
            sys.exit(1)

def type_converter(interface_comparator,interface_compared,interface_Type):
    if interface_Type == "Ethernet":
        for key in interface_comparator:
            if isinstance(interface_comparator[key], str) and isinstance(interface_compared[key], int):
                #print(type(interface_compared[key]))
                interface_compared[key] = str(interface_compared[key])
                #print(type(interface_compared[key]))
            elif isinstance(interface_comparator[key], int) and isinstance(interface_compared[key], str):
                #print(type(interface_compared[key]))
                interface_compared[key] = int(interface_compared[key])
                #print(type(interface_compared[key]))
        return interface_compared
    elif interface_Type == "Eth-Trunk":
        if interface_comparator is None:
            print("yos")
        for key, value in interface_comparator.items():
            if isinstance(value , dict) and isinstance(interface_compared[key], dict):
                for sub_key, sub_value in value.items():
                    #print(sub_value)
                    if isinstance(sub_value, str) and isinstance(interface_compared[key].get(sub_key), int):
                        print("THIS HAS BEEN CHANGED TO STR",interface_compared[key][sub_key])
                        interface_compared[key][sub_key] = str(interface_compared[key][sub_key])
                    elif isinstance(sub_value, int) and isinstance(interface_compared[key].get(sub_key), str):
                        print("THIS HAB BEEN CHANGED TO STR",interface_compared[key][sub_key])
                        interface_compared[key][sub_key] = int(interface_compared[key][sub_key])
            elif isinstance(value, str) and isinstance(interface_compared[key], int):
                print("THIS HAS BEEN CHANGE TO ",type(key))
                interface_compared[key] = str(interface_compared[key])
            elif isinstance(value, int) and isinstance(interface_compared[key], str):
                print("THIS HAS BEEN CHANGE TO ",type(interface_compared[key]))
                print(interface_compared[key])
                interface_compared[key] = int(interface_compared[key])
            # Handle boolean and NoneType values
            elif isinstance(interface_compared[key], (bool, type(None))):
                continue
        return interface_compared

def ethernet_value_comparator(interfaces_compared,interfaces_comparator,interface_Type):
        difference_counter = 0
        if interface_Type == "Ethernet":
            for key in interfaces_comparator:
                if key == "description":
                # Get description values and strip whitespace
                    comp_description = interfaces_comparator.get(key)
                    comped_description = interfaces_compared.get(key)                    
                     # Continue if both descriptions are empty
                    if comp_description is not None and comped_description is not None:
                        comp_description = comp_description.strip()
                        comped_description = comped_description.strip()
                        if comp_description != comped_description:
                            print(f"{key}: {interfaces_comparator[key]} -----> {interfaces_compared[key]}")
                        elif type(comp_description) != type(comped_description):
                            print(f"{key}: {interfaces_comparator[key]} -----> {interfaces_compared[key]}")
                            difference_counter += 1
                        elif comp_description == comped_description == "":
                            continue
                if key in interfaces_compared and interfaces_comparator[key] != interfaces_compared[key]:
                    print(f"{key}: {interfaces_comparator[key]} -----> {interfaces_compared[key]}")
                    difference_counter +=1
                elif interfaces_comparator[key] is None and interfaces_compared.get(key) is not None:
                    print(f"{key}: {interfaces_comparator[key]} -----> {interfaces_compared[key]}")
                    difference_counter +=1
                elif interfaces_comparator[key] is not None and interfaces_compared.get(key) is None:
                    print(f"{key}: {interfaces_comparator[key]} -----> {interfaces_compared[key]}")
                    difference_counter +=1
            print("NOMBRE DE DIFFERENCE",difference_counter)
            if difference_counter == 0 :       
                return False
            else :
                return True

def ethtrunk_value_comparator(ethtrunk_compared,ethtrunk_comparator,interface_Type):
        difference_counter = 0
        if interface_Type == "Eth-Trunk":
            for key in ethtrunk_comparator:
                if key == "description":
                    # Get description values and strip whitespace
                    comparator_description = ethtrunk_comparator.get(key)
                    comparded_description = ethtrunk_compared.get(key)

                    # Continue if both descriptions are empty
                    if comparator_description is not None and comparded_description is not None:
                        comparator_description = comparator_description.strip()
                        comparded_description = comparded_description.strip()
                        if comparator_description != comparded_description:
                            print(f"{key}: {ethtrunk_comparator[key]} -----> {ethtrunk_compared[key]}")
                            difference_counter += 1
                        elif type(comparator_description) != type(comparded_description):
                            print(f"{key}: {ethtrunk_comparator[key]} -----> {ethtrunk_compared[key]}")
                            difference_counter += 1
                        elif comparator_description == comparded_description == "":
                            continue
                    else:
                        print(f"{key}: {ethtrunk_comparator[key]} -----> {ethtrunk_compared[key]}")

                elif isinstance(ethtrunk_comparator[key], dict) and isinstance(ethtrunk_compared.get(key), dict):
                    # Recursively compare nested dictionaries
                     for nested_comparator_dict in ethtrunk_comparator[key]:
                        nested_comparator_dict_key = nested_comparator_dict
                        nested_comparator_dict_value = ethtrunk_comparator[key][nested_comparator_dict]
                        for nested_compared_dict in ethtrunk_compared[key]:
                            nested_compared_dict_key = nested_compared_dict
                            nested_compared_dict_value = ethtrunk_compared[key][nested_compared_dict]
                            if nested_comparator_dict_key == nested_compared_dict_key:
                                if nested_comparator_dict_value != nested_compared_dict_value:
                                        print(f"{nested_comparator_dict_key}: {nested_comparator_dict_value} -----> {nested_compared_dict_value}")
                                        difference_counter += 1
                                elif nested_comparator_dict_value is None and nested_compared_dict_value is not None:
                                    print(f"{nested_comparator_dict_key}: {nested_comparator_dict_value} -----> {nested_compared_dict_value}")
                                    difference_counter += 1
                                elif nested_comparator_dict_value is not None and nested_compared_dict_value is None:
                                    print(f"{nested_comparator_dict_key}: {nested_comparator_dict_value} -----> {nested_compared_dict_value}")
                                    difference_counter += 1
                                elif nested_comparator_dict_value is None and nested_compared_dict_value is None:
                                    pass               
                    #print(ethtrunk_comparator[key],ethtrunk_compared.get(key))
                elif ethtrunk_comparator[key] != ethtrunk_compared.get(key):
                    print(f"{key}: {ethtrunk_comparator[key]} -----> {ethtrunk_compared[key]}")
                    difference_counter += 1
                elif ethtrunk_comparator[key] is None and ethtrunk_compared.get(key) is not None:
                    print(f"{key}: {ethtrunk_comparator[key]} -----> {ethtrunk_compared[key]}")
                    difference_counter += 1
                elif ethtrunk_comparator[key] is not None and ethtrunk_compared.get(key) is None:
                    print(f"{key}: {ethtrunk_comparator[key]} -----> {ethtrunk_compared[key]}")
                    difference_counter += 1
                elif ethtrunk_comparator[key] is None and ethtrunk_compared.get(key) is None:
                    pass

            #print("NOMBRE DE DIFFERENCE", difference_counter)
            if difference_counter == 0:
                return False
            else:
                return True

def Ethernet_Request(token, stackId,device_name,deviceId,device_interface_list,request_interface_dict,interface_Type,ethernet_interface_Name,already_seen_devices_list):
   # Initialize device_interface_dict outside of the if-else block
    if stackId is not None and len(stackId) == 36:
        if stackId not in already_seen_devices_list:
            device_interface_list = get_interfaces(token, stackId, interface_Type)
            already_seen_devices_list.append(stackId)
        else:
            pass# This line seems redundant, you can remove it
        #print("------------------------------------",device_interface_list )
        request_interface_dict = dict_comparator(request_interface_dict, device_interface_list,interface_Type)
        #interface_comparison = None
        for interfaces_in_device in device_interface_list :
            if interfaces_in_device["name"] == ethernet_interface_Name:
                interface_comparison = interfaces_in_device
        request_interface_dict = type_converter(interface_comparison, request_interface_dict, interface_Type)
        differences_found = ethernet_value_comparator(request_interface_dict, interface_comparison, interface_Type)
        if not differences_found:
            print(f"{ethernet_interface_Name}, No modification detected")
            return [differences_found, already_seen_devices_list,device_interface_list]
        else:
            #print("NESTED DICT AVANT LA REQUETE ", request_interface_dict)
            update_ethernet_request(token, request_interface_dict, stackId,device_name,ethernet_interface_Name)
            return [differences_found, already_seen_devices_list,device_interface_list]
    # Initialize device_interface_dict outside of the if-else block
    elif stackId is None or len(stackId) != 36 and deviceId is not None and len(deviceId) == 36:
        if deviceId not in already_seen_devices_list:
            device_interface_list = get_interfaces(token,deviceId,interface_Type)
            already_seen_devices_list.append(deviceId)
        else:
            pass# This line seems redundant, you can remove it
        #print("------------------------------------",device_interface_list)
        request_interface_dict = dict_comparator(request_interface_dict, device_interface_list,interface_Type)
        #interface_comparison = None
        for interfaces_in_device in device_interface_list:
            if interfaces_in_device["name"] == ethernet_interface_Name:
                interface_comparison = interfaces_in_device
        request_interface_dict = type_converter(interface_comparison, request_interface_dict, interface_Type)
        differences_found = ethernet_value_comparator(request_interface_dict, interface_comparison, interface_Type)
        if not differences_found:
            print(f"{ethernet_interface_Name}, No modification detected")
            return [differences_found, already_seen_devices_list,device_interface_list]
        else:
            #print("NESTED DICT AVANT LA REQUETE ", request_interface_dict)
            update_ethernet_request(token, request_interface_dict, deviceId,device_name,ethernet_interface_Name)
            return [differences_found, already_seen_devices_list,device_interface_list]

def EthTrunk_Request(token, stackId,device_name,deviceId,device_ethtrunk_list,request_ethtrunk_dict, interface_Type, ethtrunk_interface_Name, already_seen_devices_list):
   # Initialize device_interface_dict outside of the if-else block
    if stackId is not None and len(stackId) == 36:
        if stackId in already_seen_devices_list and interface_Type == "Eth-Trunk":
             device_ethtrunk_list = get_interfaces(token, stackId, interface_Type)
        # This line seems redundant, you can remove it
        #print("------------------------------------",device_interface_list )
        request_ethtrunk_dict = dict_comparator(request_ethtrunk_dict,device_ethtrunk_list,interface_Type)
        ethtrunk_comparison = None
        for ethtrunk_in_device in device_ethtrunk_list :
            if ethtrunk_in_device["name"] == ethtrunk_interface_Name:
                ethtrunk_comparison = ethtrunk_in_device
                #print('interface sur le switch avant modification',ethtrunk_comparison)
        request_ethtrunk_dict = type_converter(ethtrunk_comparison, request_ethtrunk_dict, interface_Type)
        differences_found = ethtrunk_value_comparator(request_ethtrunk_dict, ethtrunk_comparison, interface_Type)
        if not differences_found:
            print(f"{ethtrunk_interface_Name}, No modification detected")
            return [differences_found, already_seen_devices_list,device_ethtrunk_list]
        else:
            #print("NESTED DICT AVANT LA REQUETE ", request_ethtrunk_dict)
            update_ethtrunk_request(token, request_ethtrunk_dict, stackId,device_name,ethtrunk_interface_Name)
            return [differences_found, already_seen_devices_list,device_ethtrunk_list]
    # Initialize device_interface_dict outside of the if-else block
    elif stackId is None or len(stackId) != 36 and deviceId is not None and len(deviceId) == 36:
        if deviceId not in already_seen_devices_list:
            device_ethtrunk_list = get_interfaces(token, deviceId, interface_Type)
            already_seen_devices_list.append(deviceId)
        else:
            pass# This line seems redundant, you can remove it
        #print("------------------------------------",device_interface_list )
        request_ethtrunk_dict = dict_comparator(request_ethtrunk_dict,device_ethtrunk_list,interface_Type)
        #ethtrunk_comparison = None
        for ethtrunk_in_device in device_ethtrunk_list :
            if ethtrunk_in_device["name"] == ethtrunk_interface_Name:
                ethtrunk_comparison = ethtrunk_in_device
        request_ethtrunk_dict = type_converter(ethtrunk_comparison, request_ethtrunk_dict, interface_Type)
        differences_found = ethtrunk_value_comparator(request_ethtrunk_dict, ethtrunk_comparison, interface_Type)
        if not differences_found:
            print(f"{ethtrunk_interface_Name}, No modification detected")
            return [differences_found, already_seen_devices_list,device_ethtrunk_list]
        else:
            #print("NESTED DICT AVANT LA REQUETE ", request_ethtrunk_dict)
            update_ethtrunk_request(token, request_ethtrunk_dict, deviceId,device_name,ethtrunk_interface_Name)
            return [differences_found, already_seen_devices_list,device_ethtrunk_list]
        

def Import(token,filename_or_path):
    path = rf"{filename_or_path}"
    wb = load_workbook(filename=path)
    sheet = wb.active
    sheet_name = sheet.title


    first_row_headers = [cell.value for cell in sheet[1]]

    # Initialize an empty list to store the rows
    interfaces = []
    devices_interfaces = None
    already_seen_devices = []

    # Add each value in the worksheet to the list as a tuple with its column header
    for row in sheet.iter_rows(min_row=2, values_only=True):
        interfaces.append([(first_row_headers[i], value) for i, value in enumerate(row)])

    for interfaces_row in interfaces:
        inside_dto = False
        differences_found = False


        device_Name = next((key_value_tuple[1] for key_value_tuple in interfaces_row if key_value_tuple[0] == 'device_Name'), None)
        interface_Type = next((key_value_tuple[1] for key_value_tuple in interfaces_row if key_value_tuple[0] == 'interface_Type'), None)
        stack_Id = next((key_value_tuple[1] for key_value_tuple in interfaces_row if key_value_tuple[0] == 'stack_Id'), None)
        device_Id = next((key_value_tuple[1] for key_value_tuple in interfaces_row if key_value_tuple[0] == 'device_Id'), None)
        interface_Name_and_ethtrunk_Name = next((key_value_tuple[1] for key_value_tuple in interfaces_row if key_value_tuple[0] == 'name'), None)
        status = next((key_value_tuple[1] for key_value_tuple in interfaces_row if key_value_tuple[0] == 'status'), None)

        excluded_keys = ['device_Name', 'interface_Type','device_Id','stack_Id']  # Values you want to exclude

        interfaces_row = [(key, value) for key, value in interfaces_row if key not in excluded_keys]
        if interface_Type == "Ethernet" and status != None:
            modified_ethernet_request = {}
            for key, value in interfaces_row:
                if isinstance(value,str) and value.strip() == 'True':
                    value = True
                elif isinstance(value,str) and value.strip() == 'False':
                    value = False
                    # Convert float
                if isinstance(value,str) and value.strip() == 'None' and key != 'description':
                    value = None
                    modified_ethernet_request[key] = value
                if isinstance(value, float):
                    value = int(value)
                if isinstance(value, str) and value in ' ' and key != 'description':
                    pass
                else:
                    modified_ethernet_request[key] = value

            #print("NESTED DICT ------------",modified_ethernet_request)
            Ethernet_request_result = Ethernet_Request(token,stack_Id,device_Name,device_Id,devices_interfaces,modified_ethernet_request,interface_Type,interface_Name_and_ethtrunk_Name,already_seen_devices)
            differences_found = Ethernet_request_result[0]
            already_seen_devices = Ethernet_request_result[1]
            devices_interfaces = Ethernet_request_result[2]
            if not differences_found:
                print(f"➡️  No modification detected \nInterface: {interface_Name_and_ethtrunk_Name} \nSwitch:{device_Name} ")
                continue

            ethernet_json = json.dumps(modified_ethernet_request, indent=4)       
            # Write JSON string to a file
            with open('ethernet_request.json', 'w') as json_file:
                json_file.write(ethernet_json) 
        
        elif status == "None" and interface_Type == "Ethernet":
            print(f"{interface_Name_and_ethtrunk_Name} is a stack interface it shall not be change !")


        if interface_Type == "Eth-Trunk" and status != "None":
            modified_ethtrunk_request = {}
            dto_data = {}
            inside_dto = False
            for key, value in interfaces_row:
                # Handle "None" strings
                if isinstance(value,str) and value.lower().strip() == 'true':
                    value = True
                if isinstance(value,str) and value.lower().strip() == 'false':
                    value = False
                    # Convert float to int
                if isinstance(value, float):
                    value = int(value)
                if isinstance(value, str) and value.strip() == 'None' and key != 'description':
                    value = None
                # Skip empty string values
                if isinstance(value, str) and value.strip() in " " and key != 'description':
                    continue
                
                # Handle portMemberList
                if key == 'portMemberList':
                    words = value.split(',')
                    #print("la liste des interfaces", words)
                    if len(words) == 1:
                        modified_ethtrunk_request[key] = [words[0].strip()]
                    else:
                        modified_ethtrunk_request[key] = [word.strip() for word in words]
                                    
                # Handle LswEthTrunkLacpConfigDto
                elif key == 'LswEthTrunkLacpConfigDto':
                    modified_ethtrunk_request[key] = dto_data
                    inside_dto = True
                
                # Populate modified_ethtrunk_request with other key-value pairs
                elif inside_dto:
                    dto_data[key] = value
                else:
                    modified_ethtrunk_request[key] = value

            Ethtrunk_request_result = EthTrunk_Request(token,stack_Id,device_Name,device_Id,devices_interfaces,modified_ethtrunk_request,interface_Type,interface_Name_and_ethtrunk_Name,already_seen_devices)
            differences_found = Ethtrunk_request_result[0]
            already_seen_devices = Ethtrunk_request_result[1]
            devices_interfaces = Ethtrunk_request_result[2]
            if not differences_found:
                print(f"➡️  No modification detected \nInterface: {interface_Name_and_ethtrunk_Name} \nSwitch:{device_Name} ")
                continue
            ethtrunk_json = json.dumps(modified_ethtrunk_request, indent=4)       
            # Write JSON string to a file
            with open('ethtrunk_request.json', 'w') as json_file:
                json_file.write(ethtrunk_json) 
        

        elif status == "None" and interface_Type == "Eth-Trunk":
            print(f"{interface_Name_and_ethtrunk_Name} is a stack interface it shall not be change !")

    date_time = str(datetime.now()).replace("-","_").replace(' ',"_").replace(":","_").replace(".","_")
    wb.save(f"{sheet_name}_{date_time}.xlsx")
    print(f"Previous config saved as :{sheet_name}_{date_time}.xlsx")

def main():
    TOKEN = get_token()
    
    def Export(token,site_Name,filename):
        all_devices = get_device(token)
        sites_Devices = Device_Sorter(site_Name,all_devices)
        export_xlsx(token,sites_Devices,filename)

    if args.s and args.f:
        if args.f.endswith('.xlsx'):
            Export(TOKEN, args.s, args.f)
        else:
            print("❌ \033[0mError: The output file must have a .xlsx extension.\033[0m")
    else:
        print("❌ \033[1mError: Both site and filename options are required for the export or the import .\033[0m")

    if args.e:
        if not os.path.exists(args.e) or not os.path.isfile(args.e):
            print("❌ \033[1mError: The path provided is not valid.\033[0m")
        elif not args.e.endswith('.xlsx'):
            print("❌ \033[1mError: The file must have a .xlsx extension.\033[0m")
    else:
        print("❌ \033[1mError: Both site and filename options are required for the export or the import .\033[0m")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Process some options.')
    parser.add_argument('-s', metavar='site', type=str, help='Specify the site')
    parser.add_argument('-f', metavar='filename', type=str, help='Export data to the specified file (must have .xlsx extension)')
    parser.add_argument('-e', metavar='path', type=str, help='Specify the path')
    args = parser.parse_args()
    main()





""" t = get_device("x-9c1jqkdhphsbvu9jmkrv44o47xuq5e7zs8ftgbdimlhg5ig5hh1g6l6oo445ipen9dhjhddjapo47x4brxhcqos5uncaobvxhi86g9nzo59jjv5joa9g453s48vseo5e")
p = Device_Sorter("HSM_GL1",t)
print(p) """

""" f = [['S5735-LAB-24p-BLOSSOM', 'e8297fad-b69a-415c-a6e7-8700f0d8aa9a', '21980109274EL8001200','MAGASIN_TEST'], ['S5735-LAB-24p-SR03_1', '9b57f5b1-0fcc-47be-b280-e34cfd39dac7', '21980109274EL8001546','MAGASIN_TEST'], ['S5735-LAB-24p-SR03_2', '4e896b7a-8ebb-4ae2-9637-2dcdf0f09208', '21980109274EL8001211','MAGASIN_TEST'], ['S5735-LAB-48S-SR02-1', '9770a63e-68d7-4f04-8fa1-fb9761055cbd', '21980109472SL9500387','MAGASIN_TEST'], ['S5735-LAB-48p-SR02-2', '0f5359b9-4794-4fff-b902-b3b4d7623e7f', '21980109472SL9500375','MAGASIN_TEST'], ['S6730-LAB-S01_1', '7e13307c-e72c-44ca-b684-41b90788e772', 'DM2110006264','MAGASIN_TEST'], ['S6730-LAB-S01_2', 'cb92e63b-e5ca-494b-88bf-c3ce792015c8', 'DM2110006311','MAGASIN_TEST']]

f = export_xlsx("x-9f46fs89mmo8fv06qmarpck65f86s8dhgb3shh47dhaqbv5e3xiktcdd2o9h875ieqrsemk4lddenynwsb87g87y0ag5qpmp1ic4tc7zqpbybtpjo71emr3zvz7ujt7y", f)
""" 

"""
f = [['S5735-HSM-GL1--1A_230', 'e12d275b-1e55-4a02-b188-7d7ee83ff5f4', '21980109272SL9502164'], ['S5735-HSM-GL1--1A_59_2', 'f239c1d8-c4f7-4796-b6f5-846b042db85c', '2S2220003700'], ['S5735-HSM-GL1--1C_92_2', '698bf20a-a098-4bcb-bf48-53fd70f30cc0', '21980109272SL9502271'], ['S5735-HSM-GL1--1D_141_2', 'e8d93aea-a2c1-485c-9d03-aeac7204eb8b', '2S2220003712'], ['S5735-HSM-GL1--2E_19_2', '3243bf23-d6e3-4bef-90b2-d71a7de481bb', '21980109272SL9502184'], ['S5735-HSM-GL1--2G_181', '5e8e868f-a809-48cc-b0ca-ae25c4812195', '21980109272SL9502238'], ['S5735-HSM-GL1--2H_48_2', 'eaf5212e-eb8e-4e32-902e-8b928c870f1d', '21980109272SL9501869'], ['S5735-HSM-GL1--2H_48_3', '5aeb74f1-6659-49db-a265-15d2f58dd89d', '21980109272SL9501862'], ['S5735-HSM-GL1-0A_25_2', '69d300a2-fc9c-4803-8ede-9f72166b0e9b', '2S2220003734'], ['S5735-HSM-GL1-0B_145_2', '4219d6da-ead7-4f65-9881-b3b16b98fe50', '2S2220003724'], ['S5735-HSM-GL1-0B_145_3', 'b540772b-c21e-421e-b90d-f5d1eb809651', '2S2220003621'], ['S5735-HSM-GL1-0C_104_2', '0182561e-6434-4234-9657-df44cdc5936b', '21980109272SL9501961'], ['S5735-HSM-GL1-0E_232_2', 'bb27ac94-aca5-43c5-aa9f-f5f4fb891a9e', '21980109272SL9501889'], ['S5735-HSM-GL1-1A_150_2', '45d0e9a6-f8b6-4a8f-aa6f-b797075c6a43', '2S2220003676'], ['S5735-HSM-GL1-1B_152_2', '56269e90-914a-4646-9bb6-08e703be9768', '2S2220003722'], ['S5735-HSM-GL1-1C_90_2', 'c2ed87e1-ac16-48bf-9aef-01e0eb714fe3', '2S2220003648'], ['S5735-HSM-GL1-1C_90_3', 'b27196f4-7797-4290-9795-929304087ef6', '2S2220003634'], ['S5735-HSM-GL1-2A_155_2', '4abbf517-6f8e-4726-afba-161664aa1f84', '2S2220003719'], ['S5735-HSM-GL1-2C_158_2', 'f2384c0e-a225-4e52-84a5-73e9e23411b2', '21980109272SL9501863'], ['S5735-HSM-GL1-2_27_2', 'f2cdf739-7be0-479a-98b2-b31308b8d985', '21980109272SL9502267'], ['S5735-HSM-GL1-3A_160_2', '72b7a137-bc8b-4303-9f1b-c0b7c8cb7766', '21980109272SL9501964'], ['S5735-HSM-GL1-3BA_162_2', 'de3b2539-948c-4b5e-88c4-00484626ffc4', '2S2220003702'], ['S5735-HSM-GL1-3B_26_2', 'd0a7e2c8-8540-43ab-a6a2-220d3e6bb528', '2S2220003717'], ['S5735-HSM-GL1-3C_167_2', 'c71eda89-20c0-453e-8d9d-486cc7afec92', '21980109272SL9501208'], ['S5735-HSM-GL1-4B_14_2', '57c3aad4-6c16-41e0-a055-9a2591d415e2', '21980109272SL9502168'], ['S5735-HSM-GL1-5A_52_2', '7050186c-9b6c-4ef6-b30c-e66d20d987a8', '2S2220003673'], ['S5735-HSM-GL1-5B_178_2', '16755d0f-ae72-4e33-8649-8dd9cef297bd', '2S2220003708'], ['S5735-HSM-GL1-5C_180_2', 'a059e942-19c1-4beb-8b1a-b38661fb5c12', '2S2220003675'], ['S5735-HSM-GL1-6B_34_2', 'b8fed7ac-fbc3-462d-97b3-aaebc760566a', '2S2220003739'], ['S5735-HSM-GL1-6C_36_2', '4180ce0f-2334-4998-9d4d-6f85b6d0b053', '21980109272SL9502248'], ['S5735-HSM-GL1-6C_36_3', 'a38704db-d731-4113-bf03-47219ec822a8', '21980109272SL9502179'], ['S5735-HSM-GL1-7A_188_2', 'b5ed749e-640b-47a2-bbd4-d176ca2c7c73', '2S2220003720'], ['S5735-HSM-GL1-7B7C_21_2', 'ae2d2f6a-cab5-4424-a5f0-a5ddf0da8dea', '4E2130012956'], ['S5735-HSM-GL1-7B_43_2', '7c47c659-b985-43ed-a181-c8d563f56676', '2S2220003631'], ['S5735-HSM-GL1-7B_43_3', '432fb34a-8c87-41a4-b885-41f5a4ef010b', '2S2220003619'], ['S5735-HSM-GL1-7B_43_4', 'c3a21bba-7d52-49b9-b10c-cf2d1c396ba5', '2S2220003637'], ['S5735-HSM-GL1-7C_22_2', 'c53b3309-c8a3-47c4-a1f8-b562046de306', '2S2220003684'], ['S5735-HSM-GL1-8B_30_2', '43a58202-04ce-4b6d-b9d6-599d77689e4f', '21980109272SL9502120'], ['S5735-HSM-GL1-8B_30_3', '87ed7096-a35a-4baf-9304-e18920edbddb', '21980109272SL9502233'], ['S5735-HSM-GL1-8T_140', '7efe2ded-82f3-47ca-8116-b8f28227426a', '21980109272SL9502037'], ['S5735-HSM-GL1-9B_18_2', '19921334-5d55-412a-93db-3b1cb32280b7', '2S2220003618'], ['S5735-HSM-GL1-9B_18_3', '61750736-b747-42c3-b031-97f65459eea5', '2S2220003669'], ['S5735-HSM-GL1-R1_169_2', '988b2d44-1eca-495d-8a03-7e4c177ed28d', '21980109472SL9500368'], ['S5735-HSM-GL1-R1_169_3', '222e6e52-1b51-4098-b4d4-7d2380db7594', '21980109472SL9500492'], ['S5735-HSM-GL1-R1_169_4', '5f7f3f85-e469-45f5-8c2d-5ebf7e516929', '21980109472SL9500415'], ['S5735-HSM-GL1-R2_170', '0ff427a5-6609-4129-9379-7bc523003089', '21980109472SL9500294'], ['S5735-HSM-GL1-R2_170_3', 'cbd3ec8c-d38c-46cf-bb41-7a5c32f625d2', '21980109472SL9500503'], ['S5735-HSM-GL1-R2_170_4', 'd751370f-6e67-41e7-90b7-9764fb272a58', '21980109472SL9500374'], ['S5735-HSM-GL1-R3_171_2', 'f0850863-cd1e-4047-b6dd-a256ce3a0b4e', '21980109472SL9500307'], ['S5735-HSM-GL1-R3_171_3', '2a6444cd-5ebe-4adc-84ec-993677991650', '21980109472SL9500373'], ['S5735-HSM-GL1-R3_171_4', '774d3b3b-41c7-4ccc-9f17-028a1f25b9b3', '21980109472SL9500418'], ['S5735-HSM-GL1-R4_172_02', 'ea9a9808-1d4d-4abe-9550-4ee6cdf6bb5a', '21980109472SL9500315'], ['S5735-HSM-GL1-R4_172_03', 'e87a23a1-8454-4f5b-9130-a95d85eae701', '21980109472SL9500313'], ['S5735-HSM-GL1-R4_172_04', 'ece552cc-c822-4d95-9303-042a5aa161c9', '21980109472SL9500311'], ['S5735-HSM-GL1-R4_172_05', '4477b901-a0d7-43c0-a897-875a48ad6796', '21980109472SL9500317'], ['S5735-HSM-GL1-R4_172_06', '92210fec-6794-43a7-b420-c36c7a5f1eaf', '21980109472SL9500318'], ['S5735-HSM-GL1-R5_173_2', '12d50d52-5674-44d9-bea1-d27528e101fd', '21980109472SL9500284'], ['S5735-HSM-GL1-R5_173_3', '37d2b670-0374-40f9-bbfa-83743651c5b7', '4E2130013007'], ['S5735-HSM-GL1-R5_173_4', 'f4dd1a94-f006-41e8-83ac-76379c27a3ee', 
'4E2130012999'], ['S5735-HSM-GL1-R6_174_2', '1cd2fdc3-7fef-4b36-86bd-174a005cc02b', '4E2130013162'], ['S5735-HSM-GL1-R6_174_3', '2aa27378-2dcc-41d6-8acd-2616a05b6502', '21980109472SL9500386'], ['S5735-HSM-GL1-R6_174_4', '4fa89adb-e23a-4b31-8d90-fce33def78de', '4E2130012944'], ['S5735-HSM-GL1-RDC_168_-2I-04', 'f3765ef2-0dea-40f1-9537-324404471ff0', '21980109472SM4500282'], ['S5735-HSM-GL1-RDC_168_-2I-05', '3eec4e11-f7d7-40d0-bd25-0a49374c3455', '21980109472SM4500286'], ['S5735-HSM-GL1-RDC_168_-2I-06', '8e4bee7d-f4f7-4eea-a410-4f0690c48b50', '21980109472SM4500281'], ['S5735-HSM-GL1-RDC_168_3B-02', '591d066e-2180-414f-bd82-3267162d6773', '4E2130012705'], ['S5735-HSM-GL1-RDC_168_3B-03', '58db7ea2-15d8-4eab-a57e-3b26d0a81b8f', '21980109472SM4500354'], ['S5735-HSM-GL1-SS1_166_-2I-04', 'd1b78675-bcd9-450f-970b-806ac16e422e', '21980109472SM4500428'], ['S5735-HSM-GL1-SS1_166_-2I-05', '0cbb7c9b-be68-4731-b363-54c026b966c6', '4E2130012698'], ['S5735-HSM-GL1-SS1_166_-2I-06', '9fb46169-bd90-4ee0-aeb2-9034651e3f85', '4E2130012702'], ['S5735-HSM-GL1-SS1_166_3B-02', 'd582c981-a99e-42b2-a960-21b04643a832', '4E2130012678'], ['S5735-HSM-GL1-SS1_166_3B-03', '3eb83284-528a-45f9-aed2-98f6472653db', '4E2130012723'], ['S5735-HSM-GL1-SS2-Part_15', '9e5dc874-840a-4b0e-b4da-bd7118d3c58e', '21980109274ELA001740'], ['S5735-HSM-GL1-SS2-Part_15_2', 'c1118098-f9bb-483a-8352-693e07367638', '21980109274EL7000715'], ['S5735-HSM-GL1-SS2F_135_2', '64faf6b1-178a-4694-9511-f6f6a76f313c', '21980109272SL9501830'], ['S5735-HSM-GL1-SS2G_4_2', '683e4157-367d-4e4a-8a64-8fcda678b666', '21980109272SL9502163'], ['S6730-HSM-GL1-DistriCentralisee_2', '781118d8-5235-421d-8e51-33ee506eb3a6', '1022B6956272'], ['S6730-HSM-GL1-DistriGL1_2', '9b864bd8-0641-41c4-b746-8f9c5333d2f3', '102255658350'], ['S6730-HSM-GL1-DistriGL1_3', '3ad5e128-689c-4586-8218-2973bc41d9cf', '102255658484'], ['S6730-HSM-GL1-DistriGL1_4', '29687fd2-fceb-42b5-8a26-f679d9e42b4e', '102255658503'], ['S6730-HSM-GL1-HebergementServices_2', '5eacfe13-587f-40ef-8a11-a9fb28ba95c8', '1022B6868416'], ['S6730-HSM-GL1-NoeudReseau_2', '64dc6ef0-a5bb-4f18-adac-fa3e375e7b71', '1022B6956273']]

""" 
""" t = get_device("x-9c1jqkdhphsbvu9jmkrv44o47xuq5e7zs8ftgbdimlhg5ig5hh1g6l6oo445ipen9dhjhddjapo47x4brxhcqos5uncaobvxhi86g9nzo59jjv5joa9g453s48vseo5e")
p = Device_Sorter("HSM_GL1",t)
print(p)  """

"""
f = {
    "errcode": "0",
    "errmsg": "",
    "data": [
        {
            "deviceId": "b315d63a-2fb5-4b5a-a72e-6792ab3f2626",
            "stackId": ""
        }
    ]
}

print(len(f['data'][0]["stackId"])) """


""" f = [['S5735-LAB-24p-BLOSSOM', 'e8297fad-b69a-415c-a6e7-8700f0d8aa9a', '21980109274EL8001200'], ['S5735-LAB-24p-SR03_1', '9b57f5b1-0fcc-47be-b280-e34cfd39dac7', '21980109274EL8001546'], ['S5735-LAB-24p-SR03_2', '4e896b7a-8ebb-4ae2-9637-2dcdf0f09208', '21980109274EL8001211'], ['S5735-LAB-48S-SR02-1', '9770a63e-68d7-4f04-8fa1-fb9761055cbd', '21980109472SL9500387'], ['S5735-LAB-48p-SR02-2', '0f5359b9-4794-4fff-b902-b3b4d7623e7f', '21980109472SL9500375'], ['S6730-LAB-S01_1', '7e13307c-e72c-44ca-b684-41b90788e772', 'DM2110006264'], ['S6730-LAB-S01_2', 'cb92e63b-e5ca-494b-88bf-c3ce792015c8', 'DM2110006311']]

for i in f :
    print(i[0]) """
""" with open("devices.json","r") as devices_dict:
    devices_dict1 = json.load(devices_dict)


f = Device_Sorter("LSW","27CMagasinTest - MGT",devices_dict1)
print(f) """
